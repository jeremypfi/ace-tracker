#!/usr/bin/env python3
"""
Atlantic & Eastern Pacific Hurricane ACE Tracker
=================================================
Tracks Accumulated Cyclone Energy (ACE) for both Atlantic and Eastern Pacific
hurricane seasons with storm-by-storm historical data from 1991 onward.

Data Sources:
- Historical data: Tropycal library (HURDAT2 database)
- Current season: Tropycal with NHC best track data

Creates two separate spreadsheets:
- ACE_Tracker_Atlantic.xlsx
- ACE_Tracker_Pacific.xlsx

Each spreadsheet contains 5 tabs:
- Summary
- Current Season Storms
- Historical Storms (1991-present)
- Yearly Totals
- Discord Update (copy/paste ready for Discord)

Usage:
    python3 ace_tracker.py

Author: Built with Claude for JP
"""

import os
import logging
from datetime import datetime, timedelta, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tropycal.tracks as tracks

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# =============================================================================
# CONFIGURATION
# =============================================================================

OUTPUT_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")

BASINS = {
    'atlantic': {
        'name': 'Atlantic',
        'output_file': 'ACE_Tracker_Atlantic.xlsx',
        'tropycal_basin': 'north_atlantic',  # Tropycal basin name
        'normal_ace': 122.5,
        'noaa_thresholds': {
            'below_normal': 73,
            'near_normal_upper': 126,
            'above_normal_upper': 159,
        },
        'avg_named_storms': 14,
        'avg_hurricanes': 7,
        'avg_major_hurricanes': 3,
        'all_time_single_storm_ace': {'name': 'San Ciriaco (1899)', 'ace': 73.6},
    },
    'pacific': {
        'name': 'Eastern Pacific',
        'output_file': 'ACE_Tracker_Pacific.xlsx',
        'tropycal_basin': 'east_pacific',  # Tropycal basin name
        'normal_ace': 132.0,
        'noaa_thresholds': {
            'below_normal': 73,
            'near_normal_upper': 126,
            'above_normal_upper': 159,
        },
        'avg_named_storms': 15,
        'avg_hurricanes': 8,
        'avg_major_hurricanes': 4,
        'all_time_single_storm_ace': {'name': 'Fico (1978)', 'ace': 62.8},
    }
}

START_YEAR = 1991

# ACE Calculation Constants
SYNOPTIC_TIMES = ['0000', '0600', '1200', '1800']
ACE_STATUSES = ['TS', 'HU', 'SS']  # Tropical Storm, Hurricane, Subtropical Storm
MIN_NAMED_STORM_WIND = 34  # knots
MAX_STORMS_DISCORD = 10  # Maximum storms shown in Discord update before summarizing

# =============================================================================
# BACKUP DATA (used when network is unavailable)
# =============================================================================

BACKUP_DATA = {
    'atlantic': {
        'year': 2026,
        'storms': {},
        'yearly_totals': {
            2025: 130.8, 2024: 161.6, 2023: 146.0, 2022: 95.0, 2021: 145.0, 2020: 180.0,
            2019: 133.0, 2018: 136.4, 2017: 225.0, 2016: 155.0, 2015: 65.0,
            2014: 67.0, 2013: 36.0, 2012: 133.0, 2011: 126.0, 2010: 165.0,
            2009: 53.0, 2008: 146.0, 2007: 74.0, 2006: 79.0, 2005: 245.0,
            2004: 227.0, 2003: 176.0, 2002: 67.0, 2001: 106.0, 2000: 119.0,
            1999: 177.0, 1998: 182.0, 1997: 41.0, 1996: 166.0, 1995: 228.0,
            1994: 32.0, 1993: 39.0, 1992: 75.0, 1991: 34.0,
        }
    },
    'pacific': {
        'year': 2026,
        'storms': {},
        'yearly_totals': {
            2025: 127.3, 2024: 75.0, 2023: 117.0, 2022: 97.0, 2021: 86.0, 2020: 137.0,
            2019: 91.0, 2018: 316.0, 2017: 107.0, 2016: 156.0, 2015: 252.0,
            2014: 173.0, 2013: 83.0, 2012: 113.0, 2011: 60.0, 2010: 66.0,
            2009: 119.0, 2008: 85.0, 2007: 49.0, 2006: 155.0, 2005: 136.0,
            2004: 113.0, 2003: 61.0, 2002: 113.0, 2001: 83.0, 2000: 75.0,
            1999: 56.0, 1998: 187.0, 1997: 167.0, 1996: 97.0, 1995: 52.0,
            1994: 145.0, 1993: 79.0, 1992: 183.0, 1991: 79.0,
        }
    }
}

# =============================================================================
# HELPER: Storm category from max wind (knots)
# =============================================================================

def get_category(max_wind):
    if max_wind >= 137:
        return "Cat 5"
    elif max_wind >= 113:
        return "Cat 4"
    elif max_wind >= 96:
        return "Cat 3"
    elif max_wind >= 83:
        return "Cat 2"
    elif max_wind >= 64:
        return "Cat 1"
    elif max_wind >= 34:
        return "TS"
    else:
        return "TD"


def is_major(max_wind):
    return max_wind >= 96


def get_noaa_classification(ace, basin_key):
    thresholds = BASINS[basin_key]['noaa_thresholds']
    if ace >= thresholds['above_normal_upper']:
        return "Extremely Active"
    elif ace >= thresholds['near_normal_upper']:
        return "Above Normal"
    elif ace >= thresholds['below_normal']:
        return "Near Normal"
    else:
        return "Below Normal"


def finalize_storm(storm):
    ace = 0.0
    for wind in storm['wind_readings']:
        ace += (wind ** 2) / 10000.0
    storm['ace'] = round(ace, 4)
    storm['category'] = get_category(storm['max_wind'])
    storm['is_major'] = is_major(storm['max_wind'])
    if storm['start_date'] and storm['end_date']:
        storm['duration_days'] = (storm['end_date'] - storm['start_date']).days + 1
    else:
        storm['duration_days'] = 0
    return storm

# =============================================================================
# TROPYCAL DATA FETCHING
# =============================================================================

def _tropycal_basin_name(basin_key):
    """Get Tropycal basin name from configuration."""
    return BASINS[basin_key]['tropycal_basin']


def _extract_synoptic_winds(storm_obj):
    """Extract synoptic-time wind readings from a Tropycal Storm object.

    Returns list of wind speeds at 6-hourly synoptic times (00/06/12/18 UTC)
    for times when storm was at tropical storm strength or higher (>=34 kt).
    """
    wind_readings = []

    try:
        # Get time series data
        times = storm_obj.time
        winds = storm_obj.vmax

        # Filter for synoptic times only
        for i, time in enumerate(times):
            # Check if this is a synoptic time (hour must be 0, 6, 12, or 18)
            if time.hour in [0, 6, 12, 18]:
                wind = winds[i]
                # Only count if wind >= 34 knots (tropical storm strength)
                if wind >= MIN_NAMED_STORM_WIND:
                    wind_readings.append(int(wind))
    except Exception as e:
        logger.warning(f"Error extracting synoptic winds: {e}")

    return wind_readings


def parse_hurdat2(basin_key):
    """Fetch historical storm data using Tropycal TrackDataset.

    Replaces manual HURDAT2 parsing with Tropycal library.
    Returns list of storm dicts matching the existing data structure.
    """
    tropycal_basin = _tropycal_basin_name(basin_key)

    try:
        logger.info(f"Loading {tropycal_basin} data from Tropycal TrackDataset...")
        print(f"Loading historical data via Tropycal (basin: {tropycal_basin})...")

        # Create TrackDataset with include_btk=True to get latest seasons
        dataset = tracks.TrackDataset(
            basin=tropycal_basin,
            source='hurdat',
            include_btk=True
        )

        storms = []

        # Iterate through all years from START_YEAR to current
        current_year = datetime.now().year
        for year in range(START_YEAR, current_year + 1):
            try:
                season = dataset.get_season(year)

                # Process each storm in the season
                for storm_id in season.dict.keys():
                    try:
                        storm_obj = dataset.get_storm(storm_id)

                        # Extract storm attributes
                        storm_name = storm_obj.name.title() if storm_obj.name else 'UNNAMED'
                        storm_year = storm_obj.year

                        # Max wind while tropical/subtropical only — extratropical peaks
                        # don't count toward NHC hurricane classification (e.g. a storm
                        # peaking at 70kt as EX is not counted as a hurricane).
                        tropical_types = {'TD', 'TS', 'HU', 'SS', 'SD'}
                        if hasattr(storm_obj, 'type') and len(storm_obj.type) > 0:
                            trop_winds = [v for v, t in zip(storm_obj.vmax, storm_obj.type)
                                          if str(t) in tropical_types]
                            max_wind = int(max(trop_winds)) if trop_winds else (
                                int(max(storm_obj.vmax)) if len(storm_obj.vmax) > 0 else 0)
                        else:
                            max_wind = int(max(storm_obj.vmax)) if len(storm_obj.vmax) > 0 else 0

                        # Get start and end dates
                        start_date = storm_obj.time[0] if len(storm_obj.time) > 0 else None
                        end_date = storm_obj.time[-1] if len(storm_obj.time) > 0 else None

                        # Convert pandas Timestamp to datetime if needed
                        if start_date and hasattr(start_date, 'to_pydatetime'):
                            start_date = start_date.to_pydatetime()
                        if end_date and hasattr(end_date, 'to_pydatetime'):
                            end_date = end_date.to_pydatetime()

                        # Extract synoptic-time wind readings for ACE calculation
                        wind_readings = _extract_synoptic_winds(storm_obj)

                        # Build storm record
                        storm_record = {
                            'id': storm_id,
                            'name': storm_name,
                            'year': storm_year,
                            'max_wind': max_wind,
                            'wind_readings': wind_readings,
                            'start_date': start_date,
                            'end_date': end_date
                        }

                        # Finalize storm (calculates ACE, category, duration)
                        storms.append(finalize_storm(storm_record))

                    except Exception as e:
                        logger.warning(f"Error processing storm {storm_id}: {e}")
                        continue

            except Exception as e:
                # Season might not exist or have no data
                logger.debug(f"No data for {year}: {e}")
                continue

        logger.info(f"Successfully loaded {len(storms)} storms from Tropycal")
        print(f"  ✓ Loaded {len(storms)} storms from {START_YEAR}-present via Tropycal")
        return storms

    except Exception as e:
        logger.error(f"Error loading Tropycal data: {e}")
        print(f"  ✗ Error loading Tropycal data: {e}")
        print(f"  → Using backup data (yearly totals only)")
        return None

# =============================================================================
# CURRENT SEASON from Tropycal
# =============================================================================

def get_current_season(basin_key):
    """Fetch current season data using Tropycal.

    Uses TrackDataset with include_btk=True to get the most recent season data
    including preliminary best track data from NHC.

    Returns dict with: year, storms (name->ACE), storm_details (name->{ace, max_wind}), total
    """
    basin = BASINS[basin_key]
    tropycal_basin = _tropycal_basin_name(basin_key)
    current_year = datetime.now().year
    today = datetime.now().date()
    if basin_key == 'atlantic':
        season_start = datetime(current_year, 6, 1).date()
    else:
        season_start = datetime(current_year, 5, 15).date()
    season_end = datetime(current_year, 11, 30).date()
    in_active_season = season_start <= today <= season_end
    years_to_try = [current_year] if in_active_season else [current_year, current_year - 1]

    try:
        logger.info(f"Fetching current season data via Tropycal...")
        print(f"Fetching current season data via Tropycal...")

        # Create TrackDataset with include_btk=True for latest data
        dataset = tracks.TrackDataset(
            basin=tropycal_basin,
            source='hurdat',
            include_btk=True
        )

        # During active season only try current year; off-season also checks prior year
        for year in years_to_try:
            try:
                season = dataset.get_season(year)

                if not season.dict or len(season.dict) == 0:
                    continue

                storms = {}
                storm_details = {}

                # Process each storm in the season
                for storm_id in season.dict.keys():
                    try:
                        storm_obj = dataset.get_storm(storm_id)

                        # Get storm name
                        storm_name = storm_obj.name.title() if storm_obj.name else 'UNNAMED'

                        # Skip unnamed storms and numbered systems
                        if storm_name.upper() == 'UNNAMED':
                            continue

                        # Get ACE directly from Tropycal (if available)
                        # Tropycal calculates ACE for us
                        storm_ace = storm_obj.ace if hasattr(storm_obj, 'ace') and storm_obj.ace else 0.0

                        # Max wind while tropical/subtropical only (same logic as primary path)
                        tropical_types = {'TD', 'TS', 'HU', 'SS', 'SD'}
                        if hasattr(storm_obj, 'type') and len(storm_obj.type) > 0:
                            trop_winds = [v for v, t in zip(storm_obj.vmax, storm_obj.type)
                                          if str(t) in tropical_types]
                            max_wind = int(max(trop_winds)) if trop_winds else (
                                int(max(storm_obj.vmax)) if len(storm_obj.vmax) > 0 else 0)
                        else:
                            max_wind = int(max(storm_obj.vmax)) if len(storm_obj.vmax) > 0 else 0

                        # Skip TDs that never reached named-storm strength (e.g. Tropycal "One", "Two")
                        if max_wind < MIN_NAMED_STORM_WIND:
                            continue

                        # Store storm data
                        storms[storm_name] = storm_ace
                        storm_details[storm_name] = {
                            'ace': storm_ace,
                            'max_wind': max_wind
                        }

                    except Exception as e:
                        logger.warning(f"Error processing storm {storm_id}: {e}")
                        continue

                if storms:
                    total = round(sum(storms.values()), 4)
                    logger.info(f"Found {len(storms)} storms for {year} season (ACE: {total:.2f})")
                    print(f"  ✓ Found {len(storms)} storms for {year} season")
                    print(f"  ✓ Total ACE: {total:.2f}")

                    return {
                        'year': year,
                        'storms': storms,
                        'storm_details': storm_details,
                        'total': total,
                    }

            except Exception as e:
                logger.debug(f"No data for {year} season: {e}")
                continue

        # No named storms found
        if in_active_season:
            logger.info(f"No storms yet for {current_year} season")
            print(f"  ℹ No storms yet for {current_year} season — returning empty season")
            return {'year': current_year, 'storms': {}, 'storm_details': {}, 'total': 0.0}
        logger.info(f"No {current_year} storms found (likely off-season)")
        print(f"  ℹ No {current_year} storms found (likely off-season)")
        print(f"  → Using backup data...")
        return _backup_current(basin_key)

    except Exception as e:
        logger.error(f"Error fetching current season: {e}")
        print(f"  ✗ Error fetching current season via Tropycal: {e}")
        if in_active_season:
            print(f"  → Returning empty {current_year} season")
            return {'year': current_year, 'storms': {}, 'storm_details': {}, 'total': 0.0}
        print(f"  → Using backup data...")
        return _backup_current(basin_key)


def _backup_current(basin_key):
    backup = BACKUP_DATA[basin_key]
    storms = backup['storms']
    total = round(sum(storms.values()), 4)
    # Build storm_details from backup (no max_wind available)
    detail = {name: {'ace': ace, 'max_wind': 0} for name, ace in storms.items()}
    print(f"  ⚠ Using backup data for {backup['year']} season ({len(storms)} storms, ACE={total:.2f})")
    return {'year': backup['year'], 'storms': storms, 'storm_details': detail, 'total': total}


def build_current_storm_records(current):
    """Build storm detail records from current season data (Tropycal or backup).
    Returns list of dicts compatible with historical_storms format."""
    records = []
    storm_details = current.get('storm_details', {})
    year = current['year']

    for name, data in storm_details.items():
        max_wind = data.get('max_wind', 0)
        ace = data.get('ace', 0)
        cat = get_category(max_wind)
        is_major = max_wind >= 96

        records.append({
            'name': name,
            'year': year,
            'max_wind': max_wind,
            'ace': ace,
            'category': cat,
            'is_major': is_major,
            'duration_days': 0,  # Not available from climatlas
            'wind_readings': [],
        })
    return records

# =============================================================================
# CALCULATE YEARLY TOTALS & STATISTICS
# =============================================================================

def calculate_yearly_totals(storms):
    totals = {}
    for storm in storms:
        year = storm['year']
        if year not in totals:
            totals[year] = 0.0
        totals[year] += storm['ace']
    for year in totals:
        totals[year] = round(totals[year], 2)
    return totals


def calculate_yearly_stats(storms):
    """Calculate detailed stats per year for insights."""
    stats = {}
    for storm in storms:
        year = storm['year']
        if year not in stats:
            stats[year] = {
                'named_storms': 0,
                'hurricanes': 0,
                'major_hurricanes': 0,
                'ace': 0.0,
                'ace_leader': None,
                'ace_leader_value': 0.0,
                'longest_storm': None,
                'longest_days': 0,
            }
        s = stats[year]
        s['named_storms'] += 1
        if storm['max_wind'] >= 64:
            s['hurricanes'] += 1
        if storm['max_wind'] >= 96:
            s['major_hurricanes'] += 1
        s['ace'] += storm['ace']
        if storm['ace'] > s['ace_leader_value']:
            s['ace_leader'] = storm['name']
            s['ace_leader_value'] = storm['ace']
        if storm['duration_days'] > s['longest_days']:
            s['longest_storm'] = storm['name']
            s['longest_days'] = storm['duration_days']

    for year in stats:
        stats[year]['ace'] = round(stats[year]['ace'], 2)
    return stats


def find_similar_seasons(target_ace, yearly_totals, exclude_year=None):
    """Find the 3 historical seasons with ACE closest to the target."""
    candidates = [(y, ace) for y, ace in yearly_totals.items() if y != exclude_year]
    candidates.sort(key=lambda x: abs(x[1] - target_ace))
    return candidates[:3]


# =============================================================================
# GENERATE SEASON INSIGHTS
# =============================================================================

def generate_insights(basin_key, current, yearly_totals, historical_storms, yearly_stats):
    """Generate interesting facts and insights for the dashboard and Discord."""
    basin = BASINS[basin_key]
    insights = []
    current_ace = current['total']
    current_year = current['year']
    storms = current['storms']

    # 1. ACE Leader
    if storms:
        leader_name = max(storms, key=storms.get)
        leader_ace = storms[leader_name]
        leader_pct = (leader_ace / current_ace * 100) if current_ace > 0 else 0
        insights.append(f"🌀 ACE Leader: {leader_name} with {leader_ace:.1f} ACE ({leader_pct:.0f}% of season total)")

    # 2. NOAA classification
    classification = get_noaa_classification(current_ace, basin_key)
    insights.append(f"📊 Season Classification: {classification} (ACE: {current_ace:.1f})")

    # 3. Similar historical seasons
    similar = find_similar_seasons(current_ace, yearly_totals, exclude_year=current_year)
    if similar:
        similar_links = ", ".join(
            [f'<a href="history.html#{basin_key}-yr-{y}" class="sim-link">{y}</a> ({ace:.1f})' for y, ace in similar]
        )
        insights.append(f"📈 Most Similar Seasons: {similar_links}")

    # 4. Historical ranking
    all_years = list(yearly_totals.items())
    all_years.append((current_year, current_ace))
    all_years.sort(key=lambda x: x[1], reverse=True)
    rank = next(i + 1 for i, (y, _) in enumerate(all_years) if y == current_year)
    total_seasons = len(all_years)
    insights.append(f"🏆 Historical Rank: #{rank} of {total_seasons} seasons since {START_YEAR}")

    # 5. Comparison to normal
    normal = basin['normal_ace']
    pct_of_normal = (current_ace / normal * 100) if normal > 0 else 0
    above_below = "above" if current_ace > normal else "below"
    insights.append(f"📉 {pct_of_normal:.0f}% of normal ({above_below} the {normal:.1f} average)")

    # 6. Major hurricane count (from current storms — HURDAT2 or climatlas)
    current_storms_detail = [s for s in historical_storms if s['year'] == current_year] if historical_storms else []
    if not current_storms_detail:
        current_storms_detail = build_current_storm_records(current)

    if current_storms_detail:
        major_count = sum(1 for s in current_storms_detail if s['is_major'])
        hurricane_count = sum(1 for s in current_storms_detail if s['max_wind'] >= 64)
        avg_major = basin['avg_major_hurricanes']
        avg_hurricanes = basin['avg_hurricanes']
        insights.append(f"🌀 Hurricanes: {hurricane_count} (season avg: {avg_hurricanes})")
        insights.append(f"⚡ Major Hurricanes: {major_count} (season avg: {avg_major})")

    # 7. % of ACE from top storm
    if storms and current_ace > 0:
        leader_name = max(storms, key=storms.get)
        leader_ace = storms[leader_name]
        leader_pct = leader_ace / current_ace * 100
        if leader_pct > 30:
            insights.append(f"💪 Top-heavy season: {leader_pct:.0f}% of all ACE from just {leader_name}")

    # 8. Number of storms vs average
    num_storms = len(storms)
    avg_storms = basin['avg_named_storms']
    insights.append(f"🌊 Named Storms: {num_storms} (season avg: {avg_storms})")

    # 9. Longest storm this season (only from HURDAT2 which has date ranges)
    if historical_storms:
        hurdat_current = [s for s in historical_storms if s['year'] == current_year]
        if hurdat_current:
            longest = max(hurdat_current, key=lambda s: s['duration_days'])
            if longest['duration_days'] > 0:
                insights.append(f"⏱️ Longest Storm: {longest['name']} ({longest['duration_days']} days)")

    # 10. Compare to same date last year
    last_year = current_year - 1
    if last_year in yearly_totals:
        last_year_total = yearly_totals[last_year]
        insights.append(f"📅 Last Year ({last_year}) Final Total: {last_year_total:.1f} ACE")

    # 11. All-time single storm record comparison
    record = basin['all_time_single_storm_ace']
    if storms:
        leader_name = max(storms, key=storms.get)
        leader_ace = storms[leader_name]
        pct_of_record = leader_ace / record['ace'] * 100
        if pct_of_record > 50:
            insights.append(f"🎯 {leader_name} at {pct_of_record:.0f}% of all-time single-storm record ({record['name']}: {record['ace']})")

    return insights

# =============================================================================
# GENERATE DISCORD UPDATE TEXT
# =============================================================================

def generate_discord_text(basin_key, current, yearly_totals, insights):
    """Generate copy/paste ready Discord update text."""
    basin = BASINS[basin_key]
    current_year = current['year']
    current_ace = current['total']
    storms = current['storms']
    now = datetime.now(timezone.utc)

    lines = []
    lines.append(f"🌀 **{basin['name']} ACE Update** — {now.strftime('%B %d, %Y')}")
    lines.append("")

    # Storm list sorted by ACE descending
    if storms:
        sorted_storms = sorted(storms.items(), key=lambda x: x[1], reverse=True)

        # Show top storms
        shown = sorted_storms[:MAX_STORMS_DISCORD]
        remaining = sorted_storms[MAX_STORMS_DISCORD:]

        for name, ace in shown:
            lines.append(f"{name} = {ace:.1f}")

        if remaining:
            remaining_ace = sum(ace for _, ace in remaining)
            lines.append(f"+ {len(remaining)} other storms = {remaining_ace:.1f}")

    lines.append("")
    lines.append(f"**Total for {current_year} Season = {current_ace:.1f}**")

    # Comparison to last year (in JP's preferred format)
    last_year = current_year - 1
    if last_year in yearly_totals:
        lines.append(f"Total for {last_year} Season (Final) = {yearly_totals[last_year]:.1f}")

    # NOAA classification
    classification = get_noaa_classification(current_ace, basin_key)
    normal = basin['normal_ace']
    pct = (current_ace / normal * 100) if normal > 0 else 0
    lines.append(f"Season Status: {classification} ({pct:.0f}% of normal)")

    lines.append("")

    # Historical rank
    all_years = list(yearly_totals.items())
    all_years.append((current_year, current_ace))
    all_years.sort(key=lambda x: x[1], reverse=True)
    rank = next(i + 1 for i, (y, _) in enumerate(all_years) if y == current_year)
    total_seasons = len(all_years)
    lines.append(f"Historical Rank: #{rank} of {total_seasons} (since {START_YEAR})")

    # Top 3 similar seasons
    similar = find_similar_seasons(current_ace, yearly_totals, exclude_year=current_year)
    if similar:
        similar_str = ", ".join([f"{y} ({ace:.1f})" for y, ace in similar])
        lines.append(f"Similar Seasons: {similar_str}")

    # Fun facts (pick top 3 non-redundant insights)
    fact_insights = [i for i in insights if not any(skip in i for skip in ['Classification', 'Similar', 'Rank', 'of normal'])]
    if fact_insights:
        lines.append("")
        for fact in fact_insights[:3]:
            lines.append(fact)

    return "\n".join(lines)

# =============================================================================
# CREATE SPREADSHEET
# =============================================================================

def create_spreadsheet(basin_key, historical_storms, current, yearly_totals, yearly_stats, insights, discord_text):
    basin = BASINS[basin_key]
    current_year = current['year']
    current_ace = current['total']
    storms = current['storms']

    wb = Workbook()

    # --- Colors and styles ---
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='1a2d4a')
    subheader_font = Font(name='Arial', bold=True, size=10)
    subheader_fill = PatternFill('solid', fgColor='d6e4f0')
    data_font = Font(name='Arial', size=10)
    highlight_fill = PatternFill('solid', fgColor='FFF2CC')
    major_fill = PatternFill('solid', fgColor='FFD7D7')
    insight_fill = PatternFill('solid', fgColor='E8F5E9')
    discord_fill = PatternFill('solid', fgColor='5865F2')
    discord_header_font = Font(name='Arial', bold=True, color='FFFFFF', size=12)
    discord_text_font = Font(name='Consolas', size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header_row(ws, row, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

    def style_data_cell(ws, row, col, value, fmt=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = data_font
        cell.border = thin_border
        if fmt:
            cell.number_format = fmt
        return cell

    # =========================================================================
    # TAB 1: SUMMARY
    # =========================================================================
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.sheet_properties.tabColor = "1a2d4a"

    # Title
    ws_summary['A1'] = f"{basin['name']} Hurricane Season {current_year}"
    ws_summary['A1'].font = Font(name='Arial', bold=True, size=16, color='1a2d4a')
    ws_summary.merge_cells('A1:D1')

    ws_summary['A2'] = f"Generated: {datetime.now(timezone.utc).strftime('%B %d, %Y at %H:%M UTC')}"
    ws_summary['A2'].font = Font(name='Arial', italic=True, size=9, color='666666')

    # Key stats section
    row = 4
    stats_headers = ['Metric', 'Value']
    for col, header in enumerate(stats_headers, 1):
        ws_summary.cell(row=row, column=col, value=header)
    style_header_row(ws_summary, row, 2)

    normal = basin['normal_ace']
    pct_normal = (current_ace / normal * 100) if normal > 0 else 0
    classification = get_noaa_classification(current_ace, basin_key)

    # Rankings
    all_years = list(yearly_totals.items())
    all_years.append((current_year, current_ace))
    all_years.sort(key=lambda x: x[1], reverse=True)
    rank = next(i + 1 for i, (y, _) in enumerate(all_years) if y == current_year)
    total_seasons = len(all_years)

    stats_data = [
        ("Season ACE Total", round(current_ace, 1)),
        ("Named Storms", len(storms)),
        ("Normal Season ACE (1991-2020)", normal),
        ("% of Normal", f"{pct_normal:.0f}%"),
        ("NOAA Classification", classification),
        (f"Historical Rank (since {START_YEAR})", f"#{rank} of {total_seasons}"),
    ]

    # Add major hurricane count — use climatlas storm_details if HURDAT2 doesn't have current year
    current_detail = [s for s in historical_storms if s['year'] == current_year] if historical_storms else []
    if not current_detail:
        # Build from climatlas data
        current_detail = build_current_storm_records(current)

    if current_detail:
        major_count = sum(1 for s in current_detail if s['is_major'])
        hurricane_count = sum(1 for s in current_detail if s['max_wind'] >= 64)
        stats_data.append(("Hurricanes", hurricane_count))
        stats_data.append(("Major Hurricanes (Cat 3+)", major_count))

    # Last year comparison
    last_year = current_year - 1
    if last_year in yearly_totals:
        stats_data.append((f"{last_year} Season Total (Final)", yearly_totals[last_year]))

    for i, (metric, value) in enumerate(stats_data):
        r = row + 1 + i
        style_data_cell(ws_summary, r, 1, metric)
        style_data_cell(ws_summary, r, 2, value)

    # Similar seasons
    row_after_stats = row + 1 + len(stats_data) + 2
    ws_summary.cell(row=row_after_stats, column=1, value="Similar Historical Seasons")
    ws_summary.cell(row=row_after_stats, column=1).font = Font(name='Arial', bold=True, size=12, color='1a2d4a')

    similar = find_similar_seasons(current_ace, yearly_totals, exclude_year=current_year)
    for col, header in enumerate(['Year', 'ACE', 'Difference'], 1):
        ws_summary.cell(row=row_after_stats + 1, column=col, value=header)
    style_header_row(ws_summary, row_after_stats + 1, 3)

    for i, (year, ace) in enumerate(similar):
        r = row_after_stats + 2 + i
        style_data_cell(ws_summary, r, 1, year, '0')
        style_data_cell(ws_summary, r, 2, round(ace, 1), '0.0')
        style_data_cell(ws_summary, r, 3, round(abs(ace - current_ace), 1), '0.0')

    # Season Insights section
    insight_row = row_after_stats + 2 + len(similar) + 2
    ws_summary.cell(row=insight_row, column=1, value="Season Insights")
    ws_summary.cell(row=insight_row, column=1).font = Font(name='Arial', bold=True, size=12, color='1a2d4a')

    for i, insight in enumerate(insights):
        r = insight_row + 1 + i
        cell = ws_summary.cell(row=r, column=1, value=insight)
        cell.font = data_font
        cell.fill = insight_fill
        ws_summary.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

    # Column widths
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 15

    # =========================================================================
    # TAB 2: CURRENT SEASON STORMS
    # =========================================================================
    ws_storms = wb.create_sheet(f"{current_year} Storms")
    ws_storms.sheet_properties.tabColor = "ff6b35"

    headers = ['Storm', 'ACE', '% of Total', 'Category', 'Max Wind (kt)', 'Duration (days)']
    for col, header in enumerate(headers, 1):
        ws_storms.cell(row=1, column=col, value=header)
    style_header_row(ws_storms, 1, len(headers))

    sorted_storms = sorted(storms.items(), key=lambda x: x[1], reverse=True)

    # Try to match with historical_storms for extra detail, or use climatlas details
    storm_details = {}
    if historical_storms:
        for s in historical_storms:
            if s['year'] == current_year:
                storm_details[s['name'].title()] = s

    # If HURDAT2 doesn't have current year, build from climatlas data
    if not storm_details:
        for rec in build_current_storm_records(current):
            storm_details[rec['name']] = rec

    for i, (name, ace) in enumerate(sorted_storms):
        row = 2 + i
        pct = (ace / current_ace * 100) if current_ace > 0 else 0
        detail = storm_details.get(name, None)

        style_data_cell(ws_storms, row, 1, name)
        style_data_cell(ws_storms, row, 2, round(ace, 2), '0.00')
        style_data_cell(ws_storms, row, 3, round(pct, 1), '0.0"%"')

        if detail:
            cat_cell = style_data_cell(ws_storms, row, 4, detail['category'])
            style_data_cell(ws_storms, row, 5, detail['max_wind'], '0')
            # Duration: show N/A for current season storms from climatlas (no date data)
            dur = detail['duration_days']
            style_data_cell(ws_storms, row, 6, dur if dur > 0 else "N/A")
            if detail['is_major']:
                for c in range(1, len(headers) + 1):
                    ws_storms.cell(row=row, column=c).fill = major_fill
        else:
            style_data_cell(ws_storms, row, 4, "—")
            style_data_cell(ws_storms, row, 5, "—")
            style_data_cell(ws_storms, row, 6, "—")

    # Totals row
    total_row = 2 + len(sorted_storms)
    ws_storms.cell(row=total_row, column=1, value="TOTAL").font = Font(name='Arial', bold=True, size=10)
    ws_storms.cell(row=total_row, column=2, value=round(current_ace, 2))
    ws_storms.cell(row=total_row, column=2).font = Font(name='Arial', bold=True, size=10)
    ws_storms.cell(row=total_row, column=2).number_format = '0.00'
    ws_storms.cell(row=total_row, column=3, value="100%").font = Font(name='Arial', bold=True, size=10)

    for col in range(1, len(headers) + 1):
        ws_storms.cell(row=total_row, column=col).fill = PatternFill('solid', fgColor='D9E2F3')
        ws_storms.cell(row=total_row, column=col).border = thin_border

    ws_storms.column_dimensions['A'].width = 18
    ws_storms.column_dimensions['B'].width = 12
    ws_storms.column_dimensions['C'].width = 12
    ws_storms.column_dimensions['D'].width = 12
    ws_storms.column_dimensions['E'].width = 15
    ws_storms.column_dimensions['F'].width = 15

    # =========================================================================
    # TAB 3: HISTORICAL STORMS
    # =========================================================================
    ws_hist = wb.create_sheet("Historical Storms")
    ws_hist.sheet_properties.tabColor = "2ed573"

    if historical_storms:
        # Include current year storms from climatlas if not in HURDAT2
        has_current_year = any(s['year'] == current_year for s in historical_storms)
        all_storms = list(historical_storms)
        if not has_current_year:
            all_storms.extend(build_current_storm_records(current))

        hist_headers = ['Year', 'Name', 'Category', 'Max Wind (kt)', 'ACE', 'Duration (days)', 'Major?']
        for col, header in enumerate(hist_headers, 1):
            ws_hist.cell(row=1, column=col, value=header)
        style_header_row(ws_hist, 1, len(hist_headers))

        sorted_historical = sorted(all_storms, key=lambda s: (-s['year'], -s['ace']))
        for i, storm in enumerate(sorted_historical):
            row = 2 + i
            style_data_cell(ws_hist, row, 1, storm['year'], '0')
            style_data_cell(ws_hist, row, 2, storm['name'].title())
            style_data_cell(ws_hist, row, 3, storm['category'])
            style_data_cell(ws_hist, row, 4, storm['max_wind'], '0')
            style_data_cell(ws_hist, row, 5, round(storm['ace'], 2), '0.00')
            # Duration: show N/A for current season storms sourced from climatlas
            dur = storm['duration_days']
            style_data_cell(ws_hist, row, 6, dur if dur > 0 else "N/A")
            style_data_cell(ws_hist, row, 7, "Yes" if storm['is_major'] else "No")

            if storm['year'] == current_year:
                for c in range(1, len(hist_headers) + 1):
                    ws_hist.cell(row=row, column=c).fill = highlight_fill

        ws_hist.column_dimensions['A'].width = 10
        ws_hist.column_dimensions['B'].width = 18
        ws_hist.column_dimensions['C'].width = 12
        ws_hist.column_dimensions['D'].width = 15
        ws_hist.column_dimensions['E'].width = 12
        ws_hist.column_dimensions['F'].width = 15
        ws_hist.column_dimensions['G'].width = 10
    else:
        ws_hist.cell(row=1, column=1, value="Historical storm data unavailable (HURDAT2 download failed)")

    # =========================================================================
    # TAB 4: YEARLY TOTALS
    # =========================================================================
    ws_yearly = wb.create_sheet("Yearly Totals")
    ws_yearly.sheet_properties.tabColor = "ffc107"

    yearly_headers = ['Rank', 'Year', 'ACE', '% of Normal', 'Classification',
                      'Named Storms', 'Hurricanes', 'Major Hurricanes']
    for col, header in enumerate(yearly_headers, 1):
        ws_yearly.cell(row=1, column=col, value=header)
    style_header_row(ws_yearly, 1, len(yearly_headers))

    all_years_data = list(yearly_totals.items())
    all_years_data.append((current_year, current_ace))
    all_years_data.sort(key=lambda x: x[1], reverse=True)

    for i, (year, ace) in enumerate(all_years_data):
        row = 2 + i
        pct = (ace / basin['normal_ace'] * 100) if basin['normal_ace'] > 0 else 0
        classification = get_noaa_classification(ace, basin_key)

        style_data_cell(ws_yearly, row, 1, i + 1, '0')
        style_data_cell(ws_yearly, row, 2, year, '0')
        style_data_cell(ws_yearly, row, 3, round(ace, 1), '0.0')
        style_data_cell(ws_yearly, row, 4, f"{pct:.0f}%")
        style_data_cell(ws_yearly, row, 5, classification)

        # Add storm counts if we have stats
        if yearly_stats and year in yearly_stats:
            ys = yearly_stats[year]
            style_data_cell(ws_yearly, row, 6, ys['named_storms'], '0')
            style_data_cell(ws_yearly, row, 7, ys['hurricanes'], '0')
            style_data_cell(ws_yearly, row, 8, ys['major_hurricanes'], '0')
        elif year == current_year:
            # Use climatlas storm_details for current year
            cur_records = build_current_storm_records(current)
            style_data_cell(ws_yearly, row, 6, len(cur_records), '0')
            style_data_cell(ws_yearly, row, 7, sum(1 for s in cur_records if s['max_wind'] >= 64), '0')
            style_data_cell(ws_yearly, row, 8, sum(1 for s in cur_records if s['is_major']), '0')
        else:
            style_data_cell(ws_yearly, row, 6, "—")
            style_data_cell(ws_yearly, row, 7, "—")
            style_data_cell(ws_yearly, row, 8, "—")

        if year == current_year:
            for c in range(1, len(yearly_headers) + 1):
                ws_yearly.cell(row=row, column=c).fill = highlight_fill

    ws_yearly.column_dimensions['A'].width = 8
    ws_yearly.column_dimensions['B'].width = 10
    ws_yearly.column_dimensions['C'].width = 10
    ws_yearly.column_dimensions['D'].width = 12
    ws_yearly.column_dimensions['E'].width = 18
    ws_yearly.column_dimensions['F'].width = 14
    ws_yearly.column_dimensions['G'].width = 12
    ws_yearly.column_dimensions['H'].width = 16

    # =========================================================================
    # TAB 5: DISCORD UPDATE
    # =========================================================================
    ws_discord = wb.create_sheet("Discord Update")
    ws_discord.sheet_properties.tabColor = "5865F2"

    # Header
    ws_discord['A1'] = "Discord Update — Copy Everything Below"
    ws_discord['A1'].font = discord_header_font
    ws_discord['A1'].fill = discord_fill
    ws_discord.merge_cells('A1:D1')

    ws_discord['A2'] = "Select cells A4 through the end, then copy and paste into Discord"
    ws_discord['A2'].font = Font(name='Arial', italic=True, size=9, color='666666')

    # Discord text content — one line per row for easy copy/paste
    discord_lines = discord_text.split('\n')
    for i, line in enumerate(discord_lines):
        row = 4 + i
        cell = ws_discord.cell(row=row, column=1, value=line)
        cell.font = discord_text_font

    ws_discord.column_dimensions['A'].width = 80

    return wb

# =============================================================================
# CONSOLE REPORT
# =============================================================================

def generate_console_report(basin_key, current, yearly_totals, insights):
    basin = BASINS[basin_key]
    current_year = current['year']
    current_ace = current['total']
    storms = current['storms']

    lines = []
    lines.append(f"\n{'─' * 50}")
    lines.append(f"  {basin['name']} Season {current_year} — ACE: {current_ace:.1f}")
    lines.append(f"{'─' * 50}")

    if storms:
        sorted_storms = sorted(storms.items(), key=lambda x: x[1], reverse=True)
        lines.append(f"\n  {'Storm':<18} {'ACE':>8}  {'% of Total':>10}")
        lines.append(f"  {'─' * 40}")
        for name, ace in sorted_storms:
            pct = (ace / current_ace * 100) if current_ace > 0 else 0
            lines.append(f"  {name:<18} {ace:>8.2f}  {pct:>9.1f}%")
        lines.append(f"  {'─' * 40}")
        lines.append(f"  {'TOTAL':<18} {current_ace:>8.2f}  {'100.0%':>10}")

    lines.append(f"\n  Season Insights:")
    for insight in insights:
        lines.append(f"    {insight}")

    return "\n".join(lines)

# =============================================================================
# PROCESS A BASIN
# =============================================================================

def process_basin(basin_key):
    basin = BASINS[basin_key]

    print("\n" + "=" * 50)
    print(f"{basin['name']} Hurricane ACE Tracker")
    print("=" * 50 + "\n")

    # Get historical data
    historical_storms = parse_hurdat2(basin_key)

    if historical_storms:
        yearly_totals = calculate_yearly_totals(historical_storms)
        yearly_stats = calculate_yearly_stats(historical_storms)
    else:
        yearly_totals = BACKUP_DATA[basin_key]['yearly_totals']
        yearly_stats = None

    # Get current season
    current = get_current_season(basin_key)

    # Generate insights
    insights = generate_insights(basin_key, current, yearly_totals, historical_storms, yearly_stats)

    # Generate discord text
    discord_text = generate_discord_text(basin_key, current, yearly_totals, insights)

    # Print console report
    print(generate_console_report(basin_key, current, yearly_totals, insights))

    # Print discord preview
    print(f"\n{'─' * 50}")
    print("  Discord Update Preview:")
    print(f"{'─' * 50}")
    print(discord_text)

    # Create spreadsheet
    print(f"\n{'─' * 50}")
    print("Creating spreadsheet...")

    wb = create_spreadsheet(basin_key, historical_storms, current, yearly_totals, yearly_stats, insights, discord_text)

    if wb:
        try:
            if not os.path.exists(OUTPUT_FOLDER):
                os.makedirs(OUTPUT_FOLDER)
                logger.info(f"Created output directory: {OUTPUT_FOLDER}")
        except OSError as e:
            logger.error(f"Failed to create directory {OUTPUT_FOLDER}: {e}")
            print(f"  ✗ Error: Could not create output folder")
            return None

        output_path = os.path.join(OUTPUT_FOLDER, basin['output_file'])
        try:
            wb.save(output_path)
            print(f"  ✓ Saved to: {output_path}")
        except (OSError, PermissionError) as e:
            logger.error(f"Failed to save {output_path}: {e}")
            print(f"  ✗ Error: Could not save {output_path}")
            return None

        if historical_storms:
            print(f"  ✓ {len(historical_storms)} storms in Historical Storms sheet")
            print(f"  ✓ {len(yearly_totals)} seasons in Yearly Totals sheet")

        print(f"  ✓ Discord Update tab ready for copy/paste")

        # Return data for dashboard
        return {
            'basin_key': basin_key,
            'current': current,
            'yearly_totals': yearly_totals,
            'yearly_stats': yearly_stats,
            'insights': insights,
        }

    return None


# =============================================================================
# DASHBOARD HTML
# =============================================================================

def _season_progress_html(basin_key, season_year):
    today = datetime.now().date()
    if basin_key == 'atlantic':
        start = datetime(season_year, 6, 1).date()
        end = datetime(season_year, 11, 30).date()
    else:
        start = datetime(season_year, 5, 15).date()
        end = datetime(season_year, 11, 30).date()
    total_days = (end - start).days + 1
    # Past season or after Nov 30 — show full completed bar
    if today > end:
        return (
            f'<div class="season-prog">'
            f'<div class="season-prog-label">Day {total_days} of {total_days} &middot; Season complete</div>'
            f'<div class="season-prog-track"><div class="season-prog-fill" style="width:100%"></div></div>'
            f'</div>'
        )
    if today < start:
        days_until = (start - today).days
        label = f'Season begins {start.strftime("%B")} {start.day} — {days_until} day{"s" if days_until != 1 else ""} away'
        return f'<div class="season-prog offseason">{label}</div>'
    day_num = (today - start).days + 1
    pct = day_num / total_days * 100
    return (
        f'<div class="season-prog">'
        f'<div class="season-prog-label">Day {day_num} of {total_days} &middot; {pct:.0f}% complete</div>'
        f'<div class="season-prog-track"><div class="season-prog-fill" style="width:{pct:.1f}%"></div></div>'
        f'</div>'
    )


def _preseason_html(basin_key, yearly_totals, current_year):
    """HTML block for the no-storms-yet state: replaces storm table + insights."""
    basin = BASINS[basin_key]
    normal = basin['normal_ace']

    totals = list(yearly_totals.values())
    avg_ace = sum(totals) / len(totals) if totals else 0
    max_year = max(yearly_totals, key=yearly_totals.get)
    min_year = min(yearly_totals, key=yearly_totals.get)
    above_count = sum(1 for v in totals if v >= 127)
    total_seasons = len(totals)
    last_year = current_year - 1
    last_ace = yearly_totals.get(last_year, 0)
    last_class = get_noaa_classification(last_ace, basin_key) if last_ace else 'N/A'

    if basin_key == 'atlantic':
        peak_note = "Activity typically peaks in August–September when Atlantic sea surface temperatures reach their annual high."
    else:
        peak_note = "The Eastern Pacific is often active earlier in the season, with storms possible as soon as May."

    facts = [
        f"📅 Last season ({last_year}): {last_ace:.1f} ACE — {last_class}",
        f"📊 Historical average: {avg_ace:.1f} ACE/season (NOAA normal: {normal})",
        f"🏆 Most active since {START_YEAR}: {max_year} ({yearly_totals[max_year]:.1f} ACE)",
        f"📉 Quietest since {START_YEAR}: {min_year} ({yearly_totals[min_year]:.1f} ACE)",
        f"🌀 {above_count} of {total_seasons} seasons since {START_YEAR} were Above Normal or stronger",
        f"☀️ {peak_note}",
    ]
    fact_items = '\n'.join(f'<li>{f}</li>' for f in facts)

    return f'''
      <div class="preseason-notice">
        <p>No named storms yet — the {current_year} season is underway but quiet so far.</p>
      </div>
      <h3>Did You Know?</h3>
      <ul class="insights">{fact_items}</ul>'''


def generate_dashboard_html(basin_data):
    """Generate a mobile-friendly HTML dashboard for both basins."""
    now = datetime.now(timezone.utc)

    def storm_rows_html(current):
        storms = current['storms']
        details = current.get('storm_details', {})
        total = current['total']
        sorted_storms = sorted(storms.items(), key=lambda x: x[1], reverse=True)
        rows = []
        for name, ace in sorted_storms:
            pct = (ace / total * 100) if total > 0 else 0
            wind = details.get(name, {}).get('max_wind', 0)
            cat = get_category(wind) if wind > 0 else '—'
            is_major = wind >= 96
            row_class = ' class="major"' if is_major else ''
            rows.append(
                f'<tr{row_class}>'
                f'<td data-v="{name}">{name}</td>'
                f'<td data-v="{ace:.6f}">{ace:.1f}</td>'
                f'<td data-v="{pct:.4f}">{pct:.1f}%</td>'
                f'<td data-v="{wind}">{cat}</td>'
                f'<td data-v="{wind}">{wind if wind > 0 else "—"}</td>'
                f'</tr>'
            )
        return '\n'.join(rows)

    def insight_items_html(insights):
        return '\n'.join(f'<li>{i}</li>' for i in insights)

    sections = []
    for bd in basin_data:
        if not bd:
            continue
        basin = BASINS[bd['basin_key']]
        current = bd['current']
        yearly_totals = bd['yearly_totals']
        insights = bd['insights']
        current_ace = current['total']
        current_year = current['year']
        normal = basin['normal_ace']
        pct_normal = (current_ace / normal * 100) if normal > 0 else 0
        classification = get_noaa_classification(current_ace, bd['basin_key'])

        details = current.get('storm_details', {})
        named = len(details)
        hurricanes = sum(1 for d in details.values() if d.get('max_wind', 0) >= 64)
        majors = sum(1 for d in details.values() if d.get('max_wind', 0) >= 96)

        preseason = not current['storms'] and current_year == datetime.now().year

        if preseason:
            lower_section = _preseason_html(bd['basin_key'], yearly_totals, current_year)
        else:
            all_years = list(yearly_totals.items()) + [(current_year, current_ace)]
            all_years.sort(key=lambda x: x[1], reverse=True)
            rank = next(i + 1 for i, (y, _) in enumerate(all_years) if y == current_year)
            total_seasons = len(all_years)
            lower_section = f'''
      <h3>Storm Breakdown</h3>
      <div class="table-wrap">
        <table>
          <thead><tr>
            <th class="sort-th" onclick="sortDash(this,0,'s')">Storm <span class="sa"></span></th>
            <th class="sort-th" onclick="sortDash(this,1,'n')">ACE <span class="sa">▼</span></th>
            <th class="sort-th" onclick="sortDash(this,2,'n')">% <span class="sa"></span></th>
            <th class="sort-th" onclick="sortDash(this,3,'n')">Category <span class="sa"></span></th>
            <th class="sort-th" onclick="sortDash(this,4,'n')">Wind (kt) <span class="sa"></span></th>
          </tr></thead>
          <tbody id="storm-{bd['basin_key']}">
            {storm_rows_html(current)}
          </tbody>
          <tfoot>
            <tr class="total-row"><td><b>TOTAL</b></td><td><b>{current_ace:.1f}</b></td><td><b>100%</b></td><td></td><td></td></tr>
          </tfoot>
        </table>
      </div>

      <h3>Season Insights</h3>
      <ul class="insights">{insight_items_html(insights)}</ul>'''

        gauge_pct = min(pct_normal, 200)

        if preseason:
            stats_grid = f'''
      <div class="stats-grid">
        <div class="stat-box ace-total">
          <div class="stat-label">Season ACE</div>
          <div class="stat-value">0.0</div>
          <div class="stat-sub">Season underway — no storms yet</div>
          <div class="gauge"><div class="gauge-fill" style="width:0%"></div></div>
        </div>
        <div class="stat-box"><div class="stat-label">Named Storms</div><div class="stat-value">0</div></div>
        <div class="stat-box"><div class="stat-label">Hurricanes</div><div class="stat-value">0</div></div>
        <div class="stat-box major-box"><div class="stat-label">Major Hurricanes</div><div class="stat-value">0</div></div>
      </div>'''
        else:
            all_years = list(yearly_totals.items()) + [(current_year, current_ace)]
            all_years.sort(key=lambda x: x[1], reverse=True)
            rank = next(i + 1 for i, (y, _) in enumerate(all_years) if y == current_year)
            total_seasons = len(all_years)
            stats_grid = f'''
      <div class="stats-grid">
        <div class="stat-box ace-total">
          <div class="stat-label">Season ACE</div>
          <div class="stat-value">{current_ace:.1f}</div>
          <div class="stat-sub">{pct_normal:.0f}% of normal ({normal})</div>
          <div class="gauge"><div class="gauge-fill" style="width:{gauge_pct/2}%"></div></div>
        </div>
        <div class="stat-box"><div class="stat-label">Classification</div><div class="stat-value small">{classification}</div></div>
        <div class="stat-box"><div class="stat-label">Named Storms</div><div class="stat-value">{named}</div></div>
        <div class="stat-box"><div class="stat-label">Hurricanes</div><div class="stat-value">{hurricanes}</div></div>
        <div class="stat-box major-box"><div class="stat-label">Major Hurricanes</div><div class="stat-value">{majors}</div></div>
        <div class="stat-box"><div class="stat-label">Rank (since {START_YEAR})</div><div class="stat-value">#{rank}<span class="stat-sub"> of {total_seasons}</span></div></div>
      </div>'''

        sections.append(f'''
    <div class="basin-card" id="{bd['basin_key']}">
      <h2>{basin['name']} — {current_year} Season</h2>
      {_season_progress_html(bd['basin_key'], current_year)}
      {stats_grid}
      {lower_section}
    </div>''')

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta name="description" content="Track the current Atlantic and Eastern Pacific hurricane season ACE (Accumulated Cyclone Energy) in real time. Updated every 6 hours during hurricane season.">
<meta name="theme-color" content="#4fc3f7">
<link rel="canonical" href="https://aceofcanes.com/">
<meta property="og:type" content="website">
<meta property="og:site_name" content="ACE Tracker">
<meta property="og:url" content="https://aceofcanes.com/">
<meta property="og:title" content="Hurricane ACE Dashboard | aceofcanes.com">
<meta property="og:description" content="Track Accumulated Cyclone Energy (ACE) for the Atlantic and Eastern Pacific hurricane seasons in real time. Updated every 6 hours from official NOAA data.">
<meta name="twitter:card" content="summary">
<meta name="twitter:title" content="Hurricane ACE Dashboard | aceofcanes.com">
<meta name="twitter:description" content="Track Accumulated Cyclone Energy (ACE) for the Atlantic and Eastern Pacific hurricane seasons in real time. Updated every 6 hours from official NOAA data.">
<link rel="icon" type="image/png" href="ace.png">
<title>Hurricane ACE Dashboard | aceofcanes.com</title>
<script>(function(){{try{{var t=localStorage.getItem('ace-theme');if(t==='light')document.documentElement.setAttribute('data-theme','light');}}catch(e){{}}}})();</script>
<style>
  :root {{
    --bg:#0a1628; --card:#132238; --box:#1a2d4a; --accent:#4fc3f7; --accent2:#29b6f6;
    --accent-h3:#81d4fa; --text:#e0e6ed; --text-strong:#ffffff; --muted:#78909c;
    --muted-dark:#546e7a; --border:#1e3a5f; --danger:#ef5350; --danger-bg:#2a1a1a;
    --danger-text:#ef8a80; --total-row:#1a2d4a; --sources-bg:#0d1b2a; --gauge-bg:#1e3a5f;
  }}
  [data-theme="light"] {{
    --bg:#f0f4f8; --card:#ffffff; --box:#e8f0fe; --accent:#0277bd; --accent2:#0288d1;
    --accent-h3:#01579b; --text:#1a2d4a; --text-strong:#0a1628; --muted:#607d8b;
    --muted-dark:#455a64; --border:#b0bec5; --danger:#d32f2f; --danger-bg:#ffeaea;
    --danger-text:#c62828; --total-row:#e8f0fe; --sources-bg:#e2ecf7; --gauge-bg:#c9daf8;
  }}
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif; background:var(--bg); color:var(--text); padding:12px; transition:background 0.2s,color 0.2s; }}
  .header {{ display:grid; grid-template-columns:1fr auto 1fr; align-items:center; margin:8px 0; padding:0 4px; }}
  h1 {{ grid-column:2; color:var(--accent); font-size:1.4em; text-align:center; display:flex; align-items:center; justify-content:center; gap:8px; }}
  .logo {{ height:1.5em; width:auto; vertical-align:middle; }}
  .theme-btn {{ grid-column:3; justify-self:end; background:transparent; border:1px solid var(--accent); color:var(--accent); border-radius:20px; padding:4px 10px; cursor:pointer; font-size:0.9em; }}
  .updated {{ text-align:center; color:var(--muted); font-size:0.8em; margin-bottom:8px; }}
  .nav-link {{ text-align:center; margin-bottom:12px; }}
  .nav-link a {{ color:var(--accent); text-decoration:none; font-size:0.85em; border:1px solid var(--accent); border-radius:20px; padding:4px 14px; }}
  .nav-link a:hover {{ background:var(--accent); color:var(--bg); }}
  .ace-explain {{ background:var(--box); border-radius:8px; padding:10px 14px; margin-bottom:14px; font-size:0.85em; }}
  .ace-explain summary {{ color:var(--accent); cursor:pointer; list-style:none; display:flex; align-items:center; gap:6px; min-height:44px; }}
  .ace-explain summary::-webkit-details-marker {{ display:none; }}
  .ace-explain summary::before {{ content:'ℹ'; font-size:1.1em; }}
  .ace-explain-hint {{ color:var(--muted); font-size:0.85em; }}
  .ace-explain p {{ color:var(--text); line-height:1.6; margin-top:8px; padding-top:8px; border-top:1px solid var(--border); }}
  .toggle {{ display:flex; justify-content:center; gap:8px; margin-bottom:16px; }}
  .toggle button {{ padding:8px 20px; border:1px solid var(--accent); background:transparent; color:var(--accent); border-radius:20px; cursor:pointer; font-size:0.9em; }}
  .toggle button.active {{ background:var(--accent); color:var(--bg); font-weight:bold; }}
  .basin-card {{ background:var(--card); border-radius:12px; padding:16px; margin-bottom:16px; display:none; }}
  .basin-card.active {{ display:block; }}
  h2 {{ color:var(--accent); font-size:1.2em; margin-bottom:12px; border-bottom:1px solid var(--border); padding-bottom:8px; }}
  h3 {{ color:var(--accent-h3); font-size:1em; margin:16px 0 8px; }}
  .stats-grid {{ display:grid; grid-template-columns:repeat(3,1fr); gap:8px; }}
  .stat-box {{ background:var(--box); border-radius:8px; padding:10px; text-align:center; }}
  .stat-box.ace-total {{ grid-column:span 3; }}
  .stat-label {{ color:var(--muted); font-size:0.75em; text-transform:uppercase; }}
  .stat-value {{ color:var(--text-strong); font-size:1.5em; font-weight:bold; }}
  .stat-value.small {{ font-size:1.1em; }}
  .stat-sub {{ color:var(--muted); font-size:0.75em; }}
  .major-box {{ border:1px solid var(--danger); }}
  .major-box .stat-value {{ color:var(--danger); }}
  .gauge {{ height:6px; background:var(--gauge-bg); border-radius:3px; margin-top:6px; }}
  .gauge-fill {{ height:100%; background:linear-gradient(90deg,var(--accent),var(--accent2),var(--danger)); border-radius:3px; transition:width 0.5s; }}
  .table-wrap {{ overflow-x:auto; background: linear-gradient(to right,var(--card) 20px,transparent 20px) left/20px 100%, linear-gradient(to left,var(--card) 20px,transparent 20px) right/20px 100%, linear-gradient(to right,rgba(0,0,0,0.18),transparent) left/16px 100%, linear-gradient(to left,rgba(0,0,0,0.18),transparent) right/16px 100%; background-repeat:no-repeat; background-attachment:local,local,scroll,scroll; }}
  table {{ width:100%; border-collapse:collapse; font-size:0.85em; }}
  th {{ background:var(--box); color:var(--accent); padding:8px 6px; text-align:left; position:sticky; top:0; }}
  th.sort-th {{ cursor:pointer; user-select:none; padding:10px 6px; }}
  th.sort-th:hover {{ color:var(--text-strong); }}
  .sa {{ font-size:0.7em; margin-left:2px; opacity:0.7; }}
  td {{ padding:6px; border-bottom:1px solid var(--border); color:var(--text); }}
  tr.major {{ background:var(--danger-bg); }}
  tr.major td {{ color:var(--danger-text); font-weight:bold; }}
  tr.total-row {{ background:var(--total-row); }}
  .insights {{ list-style:none; padding:0; }}
  .insights li {{ background:var(--box); padding:8px 10px; margin:4px 0; border-radius:6px; font-size:0.85em; border-left:3px solid var(--accent); color:var(--text); }}
  .sources {{ background:var(--sources-bg); border-top:1px solid var(--border); margin-top:24px; padding:16px 12px; border-radius:8px; }}
  .sources h4 {{ color:var(--muted); font-size:0.8em; text-transform:uppercase; margin-bottom:8px; }}
  .sources a {{ color:var(--accent); text-decoration:none; font-size:0.78em; }}
  .sources a:hover {{ text-decoration:underline; }}
  .sources p {{ color:var(--muted-dark); font-size:0.75em; margin-top:8px; line-height:1.5; }}
  .sources ul {{ list-style:none; padding:0; margin:0; }}
  .sources li {{ color:var(--muted); font-size:0.78em; margin:4px 0; padding-left:12px; position:relative; }}
  .sources li::before {{ content:"•"; position:absolute; left:0; color:var(--accent); }}
  .sources code {{ font-size:0.9em; background:var(--box); padding:1px 4px; border-radius:3px; }}
  .disclaimer {{ margin-top:12px; padding:10px 12px; border-radius:6px; border-left:3px solid var(--muted); font-size:0.75em; color:var(--muted); line-height:1.5; }}
  .kofi-link {{ text-align:center; margin-top:14px; font-size:0.78em; }}
  .kofi-link a {{ color:var(--muted); text-decoration:none; }}
  .kofi-link a:hover {{ color:var(--accent); }}
  .season-prog {{ margin:-4px 0 14px; }}
  .season-prog-label {{ color:var(--muted); font-size:0.8em; margin-bottom:5px; text-align:center; }}
  .season-prog-track {{ height:6px; background:var(--gauge-bg); border-radius:3px; }}
  .season-prog-fill {{ height:100%; background:linear-gradient(90deg,var(--accent),var(--accent2)); border-radius:3px; transition:width 0.5s; }}
  .season-prog.offseason {{ color:var(--muted); font-size:0.8em; text-align:center; margin:-4px 0 14px; }}
  .preseason-notice {{ background:var(--box); border-radius:8px; padding:14px 16px; margin:12px 0; border-left:4px solid var(--accent); font-size:0.9em; color:var(--text); text-align:center; line-height:1.5; }}
  .sim-link {{ color:var(--accent); text-decoration:none; }}
  .sim-link:hover {{ text-decoration:underline; }}
  @media(min-width:768px) {{ body {{ max-width:900px; margin:0 auto; padding:24px; }} .stats-grid {{ grid-template-columns:repeat(6,1fr); }} .stat-box.ace-total {{ grid-column:span 6; }} }}
  @media(min-width:1100px) {{ body {{ max-width:1100px; }} }}
</style>
</head>
<body>
<div class="header">
  <h1><img src="ace.png" class="logo" alt="ACE"> Hurricane ACE Dashboard</h1>
  <button class="theme-btn" id="themeBtn" onclick="toggleTheme()">☀</button>
</div>
<div class="updated">Updated: {now.strftime('%B %d, %Y at %H:%M UTC')}</div>
<div class="nav-link"><a href="history.html">📊 Season History ({START_YEAR}–present)</a></div>
<details class="ace-explain">
  <summary>What is ACE? <span class="ace-explain-hint">(tap to expand)</span></summary>
  <p>Accumulated Cyclone Energy (ACE) measures total hurricane season activity by combining storm intensity and duration. A major hurricane that lasts two weeks contributes far more than a brief tropical storm. NOAA uses seasonal ACE totals to classify years as <b>Below Normal</b> (&lt;73), <b>Near Normal</b> (73–126), <b>Above Normal</b> (126–159), or <b>Extremely Active</b> (159+).</p>
</details>
<div class="toggle">
  <button class="active" onclick="show('atlantic',this)">Atlantic</button>
  <button onclick="show('pacific',this)">Eastern Pacific</button>
</div>
{''.join(sections)}
<div class="sources">
  <h4>Data Sources</h4>
  <ul>
    <li><a href="https://www.nhc.noaa.gov/data/#hurdat" target="_blank" rel="noopener noreferrer">NOAA HURDAT2</a> — Historical best-track data (1991–present) for storm tracks, wind speeds, and ACE calculations</li>
    <li><a href="https://www.nhc.noaa.gov/data/#hurdat" target="_blank" rel="noopener noreferrer">NHC Real-time Best Track</a> — Current season preliminary storm data fetched via Tropycal (<code>include_btk=True</code>); updated continuously during active storms</li>
    <li><a href="https://www.cpc.ncep.noaa.gov/products/outlooks/background_information.shtml" target="_blank" rel="noopener noreferrer">NOAA CPC</a> — Season classification thresholds and 1991–2020 climatological normals</li>
  </ul>
  <p>ACE (Accumulated Cyclone Energy) is calculated at 6-hourly synoptic times (0000/0600/1200/1800 UTC) for systems at tropical storm strength or higher (≥34 kt), including subtropical storms. Formula: ACE = Σ(V²<sub>max</sub>) × 10⁻⁴. Categories use the Saffir-Simpson scale in knots.</p>
  <p class="disclaimer">⚠️ This site is maintained by a hurricane data enthusiast — not a meteorologist, forecaster, or weather professional of any kind. I just love the data. All information is sourced directly from official NOAA/NHC databases. For official forecasts, watches, warnings, and life-safety information, always refer to the <a href="https://www.nhc.noaa.gov/" target="_blank" rel="noopener noreferrer">National Hurricane Center</a>.</p>
  <p class="kofi-link"><a href="https://ko-fi.com/aceofcanes" target="_blank" rel="noopener noreferrer">☕ Support this project on Ko-fi</a></p>
</div>
<script>
document.querySelectorAll('.basin-card')[0]?.classList.add('active');
function show(id,btn) {{
  document.querySelectorAll('.basin-card').forEach(c=>c.classList.remove('active'));
  document.querySelectorAll('.toggle button').forEach(b=>b.classList.remove('active'));
  document.getElementById(id)?.classList.add('active');
  btn.classList.add('active');
  try{{history.replaceState(null,'','#'+id);}}catch(e){{}}
}}
function toggleTheme() {{
  var h=document.documentElement;
  var light=h.getAttribute('data-theme')==='light';
  h.setAttribute('data-theme',light?'dark':'light');
  try{{localStorage.setItem('ace-theme',light?'dark':'light');}}catch(e){{}}
  document.getElementById('themeBtn').textContent=light?'☀':'☾';
}}
document.addEventListener('DOMContentLoaded',function() {{
  document.getElementById('themeBtn').textContent=document.documentElement.getAttribute('data-theme')==='light'?'☾':'☀';
  var hash=location.hash.replace('#','');
  var match=[].slice.call(document.querySelectorAll('.toggle button')).filter(function(b){{return(b.getAttribute('onclick')||'').indexOf("'"+hash+"'")>=0;}})[0];
  if(match)match.click();
}});
var _ds={{}};
function sortDash(th,col,type){{
  var card=th.closest('.basin-card');
  var tbody=card.querySelector('tbody');
  var key=card.id+col;
  var asc=_ds[key]===undefined?false:!_ds[key];
  _ds[key]=asc;
  var rows=Array.from(tbody.querySelectorAll('tr'));
  rows.sort(function(a,b){{
    var av=a.cells[col]?a.cells[col].getAttribute('data-v'):'';
    var bv=b.cells[col]?b.cells[col].getAttribute('data-v'):'';
    if(type==='n'){{av=parseFloat(av)||0;bv=parseFloat(bv)||0;}}
    if(av<bv)return asc?-1:1;
    if(av>bv)return asc?1:-1;
    return 0;
  }});
  rows.forEach(r=>tbody.appendChild(r));
  card.querySelectorAll('.sort-th .sa').forEach((s,i)=>{{s.textContent=i===col?(asc?'▲':'▼'):'';}});
}}
</script>
</body>
</html>'''
    return html


def generate_history_html(basin_data):
    """Generate a historical seasons summary page (all seasons since START_YEAR)."""
    now = datetime.now(timezone.utc)

    def _badge_class(ace_val, basin_key):
        c = get_noaa_classification(ace_val, basin_key)
        if 'Extreme' in c:
            return 'extreme', c
        if 'Above' in c:
            return 'above', c
        if 'Below' in c:
            return 'below', c
        return 'near', c

    basin_sections = []
    for bd in basin_data:
        if not bd:
            continue
        basin = BASINS[bd['basin_key']]
        current = bd['current']
        yearly_totals = bd['yearly_totals']
        yearly_stats = bd.get('yearly_stats')
        current_year = current['year']
        normal = basin['normal_ace']

        # Build per-year data from yearly_stats (HURDAT2 historical)
        years_data = {}
        if yearly_stats:
            for year, stats in yearly_stats.items():
                years_data[year] = {
                    'ace': round(stats['ace'], 1),
                    'named': stats['named_storms'],
                    'hurricanes': stats['hurricanes'],
                    'majors': stats['major_hurricanes'],
                    'leader': stats.get('ace_leader') or '—',
                }
        else:
            for year, ace in yearly_totals.items():
                years_data[year] = {
                    'ace': round(ace, 1),
                    'named': '—', 'hurricanes': '—', 'majors': '—', 'leader': '—',
                }

        # Override current year with live data (more up-to-date than HURDAT2)
        details = current.get('storm_details', {})
        current_storms = current.get('storms', {})
        current_ace = round(current['total'], 1)
        named = len(details)
        hurricanes = sum(1 for d in details.values() if d.get('max_wind', 0) >= 64)
        majors = sum(1 for d in details.values() if d.get('max_wind', 0) >= 96)
        leader = max(current_storms, key=current_storms.get) if current_storms else '—'
        # Add current year row if season is active or has storm activity
        today = datetime.now().date()
        if bd['basin_key'] == 'atlantic':
            _season_start = datetime(current_year, 6, 1).date()
        else:
            _season_start = datetime(current_year, 5, 15).date()
        _season_end = datetime(current_year, 11, 30).date()
        _in_active_season = _season_start <= today <= _season_end and current_year == datetime.now().year
        if current_ace > 0 or named > 0 or _in_active_season:
            years_data[current_year] = {
                'ace': current_ace,
                'named': named,
                'hurricanes': hurricanes,
                'majors': majors,
                'leader': leader,
                'active': True,
            }

        # Compute ACE rank and top-5
        ranked = sorted(years_data.items(), key=lambda x: x[1]['ace'], reverse=True)
        ranks = {year: i + 1 for i, (year, _) in enumerate(ranked)}
        top5_years = {year for year, _ in ranked[:5]}
        total_seasons = len(years_data)
        max_ace = max(d['ace'] for d in years_data.values()) if years_data else 1

        # Average row values (using official NOAA 1991-2020 normals from BASINS config)
        avg_ace = normal
        avg_pct = 100
        avg_named = basin['avg_named_storms']
        avg_hurr = basin['avg_hurricanes']
        avg_major = basin['avg_major_hurricanes']

        # Classification sort key helper
        def _csort(bc):
            return {'below': 0, 'near': 1, 'above': 2, 'extreme': 3}.get(bc, 1)

        # Build table rows (year descending default)
        rows = []
        for year in sorted(years_data.keys(), reverse=True):
            d = years_data[year]
            ace = d['ace']
            pct = round(ace / normal * 100) if normal > 0 else 0
            bc, classification = _badge_class(ace, bd['basin_key'])
            rank = ranks[year]
            is_active = d.get('active', False)
            is_top5 = year in top5_years
            ace_bar_pct = round(ace / max_ace * 100, 1)
            row_cls = f'row-{bc}'
            if is_active:
                row_cls += ' row-current'
            if is_top5:
                row_cls += ' row-top5'
            active_label = ' <span class="active-dot" title="Season in progress">●</span>' if is_active else ''
            named_v = d['named'] if d['named'] != '—' else 0
            hurr_v = d['hurricanes'] if d['hurricanes'] != '—' else 0
            major_v = d['majors'] if d['majors'] != '—' else 0
            rows.append(
                f'<tr class="{row_cls}" id="{bd["basin_key"]}-yr-{year}">'
                f'<td data-v="{year}"><b>{year}</b>{active_label}</td>'
                f'<td data-v="{ace:.4f}"><b>{ace:.1f}</b><div class="ace-bar"><div class="ace-bar-fill" style="width:{ace_bar_pct}%"></div></div></td>'
                f'<td data-v="{pct}">{pct}%</td>'
                f'<td data-v="{_csort(bc)}"><span class="badge badge-{bc}">{classification}</span></td>'
                f'<td data-v="{named_v}">{d["named"]}</td>'
                f'<td data-v="{hurr_v}">{d["hurricanes"]}</td>'
                f'<td data-v="{major_v}">{d["majors"]}</td>'
                f'<td data-v="{d["leader"]}">{d["leader"]}</td>'
                f'<td data-v="{rank}">#{rank}&nbsp;/&nbsp;{total_seasons}</td>'
                f'</tr>'
            )

        # Average row (goes in tfoot, not sorted)
        avg_bar_pct = round(avg_ace / max_ace * 100, 1)
        avg_row = (
            f'<tr class="row-avg">'
            f'<td>Avg (1991–2020)</td>'
            f'<td>{avg_ace:.1f}<div class="ace-bar"><div class="ace-bar-fill" style="width:{avg_bar_pct}%"></div></div></td>'
            f'<td>{avg_pct}%</td>'
            f'<td><span class="badge badge-near">Near Normal</span></td>'
            f'<td>{avg_named}</td>'
            f'<td>{avg_hurr}</td>'
            f'<td>{avg_major}</td>'
            f'<td>—</td>'
            f'<td>—</td>'
            f'</tr>'
        )

        basin_sections.append(f'''
    <div class="basin-card" id="{bd['basin_key']}">
      <h2>{basin['name']} — All Seasons ({START_YEAR}–{current_year})</h2>
      <p class="season-note">{total_seasons} seasons &nbsp;·&nbsp; ● = currently active &nbsp;·&nbsp; <span style="border-left:3px solid #f9a825;padding-left:4px;">gold border</span> = top 5 all-time ACE &nbsp;·&nbsp; click headers to sort</p>
      <div class="table-wrap">
        <table class="hist-table">
          <thead>
            <tr>
              <th class="sort-th" onclick="sortHist(this,0,'n')">Year <span class="sa">▼</span></th>
              <th class="sort-th" onclick="sortHist(this,1,'n')">ACE <span class="sa"></span></th>
              <th class="sort-th" onclick="sortHist(this,2,'n')">% Normal <span class="sa"></span></th>
              <th class="sort-th" onclick="sortHist(this,3,'n')">Classification <span class="sa"></span></th>
              <th class="sort-th" onclick="sortHist(this,4,'n')">Named <span class="sa"></span></th>
              <th class="sort-th" onclick="sortHist(this,5,'n')">Hurr. <span class="sa"></span></th>
              <th class="sort-th" onclick="sortHist(this,6,'n')">Major <span class="sa"></span></th>
              <th class="sort-th" onclick="sortHist(this,7,'s')">ACE Leader <span class="sa"></span></th>
              <th class="sort-th" onclick="sortHist(this,8,'n')">Rank <span class="sa"></span></th>
            </tr>
          </thead>
          <tbody id="hist-{bd['basin_key']}">{''.join(rows)}</tbody>
          <tfoot>{avg_row}</tfoot>
        </table>
      </div>
    </div>''')

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta name="description" content="Compare every Atlantic and Eastern Pacific hurricane season from 1991 to present by ACE, storm counts, and NOAA activity classifications.">
<meta name="theme-color" content="#4fc3f7">
<link rel="canonical" href="https://aceofcanes.com/history.html">
<meta property="og:type" content="website">
<meta property="og:site_name" content="ACE Tracker">
<meta property="og:url" content="https://aceofcanes.com/history.html">
<meta property="og:title" content="Season History (1991–present) | aceofcanes.com">
<meta property="og:description" content="Compare every Atlantic and Eastern Pacific hurricane season from 1991 to present by ACE, storm counts, and NOAA activity classifications.">
<meta name="twitter:card" content="summary">
<meta name="twitter:title" content="Season History (1991–present) | aceofcanes.com">
<meta name="twitter:description" content="Compare every Atlantic and Eastern Pacific hurricane season from 1991 to present by ACE, storm counts, and NOAA activity classifications.">
<link rel="icon" type="image/png" href="ace.png">
<title>Season History (1991–present) | aceofcanes.com</title>
<script>(function(){{try{{var t=localStorage.getItem('ace-theme');if(t==='light')document.documentElement.setAttribute('data-theme','light');}}catch(e){{}}}})();</script>
<style>
  :root {{
    --bg:#0a1628; --card:#132238; --box:#1a2d4a; --accent:#4fc3f7;
    --text:#e0e6ed; --text-strong:#ffffff; --muted:#78909c; --border:#1e3a5f;
    --sources-bg:#0d1b2a; --gauge-bg:#1e3a5f;
    --row-extreme:rgba(239,83,80,0.10); --row-above:rgba(255,143,0,0.10);
    --row-below:rgba(66,165,245,0.10); --row-near:transparent;
    --current-border:#4fc3f7; --active-dot:#4fc3f7;
    --badge-extreme:#ef5350; --badge-above:#ff8f00; --badge-near:#546e7a; --badge-below:#1976d2;
  }}
  [data-theme="light"] {{
    --bg:#f0f4f8; --card:#ffffff; --box:#e8f0fe; --accent:#0277bd;
    --text:#1a2d4a; --text-strong:#0a1628; --muted:#607d8b; --border:#b0bec5;
    --sources-bg:#e2ecf7; --gauge-bg:#c9daf8;
    --row-extreme:rgba(198,40,40,0.07); --row-above:rgba(230,81,0,0.07);
    --row-below:rgba(21,101,192,0.07); --row-near:transparent;
    --current-border:#0277bd; --active-dot:#0277bd;
    --badge-extreme:#c62828; --badge-above:#e65100; --badge-near:#546e7a; --badge-below:#1565c0;
  }}
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif; background:var(--bg); color:var(--text); padding:12px; transition:background 0.2s,color 0.2s; }}
  .header {{ display:grid; grid-template-columns:1fr auto 1fr; align-items:center; margin:8px 0; padding:0 4px; }}
  h1 {{ grid-column:2; color:var(--accent); font-size:1.4em; text-align:center; display:flex; align-items:center; justify-content:center; gap:8px; }}
  .logo {{ height:1.5em; width:auto; vertical-align:middle; }}
  .theme-btn {{ grid-column:3; justify-self:end; background:transparent; border:1px solid var(--accent); color:var(--accent); border-radius:20px; padding:4px 10px; cursor:pointer; font-size:0.9em; }}
  .updated {{ text-align:center; color:var(--muted); font-size:0.8em; margin-bottom:8px; }}
  .nav-link {{ text-align:center; margin-bottom:12px; }}
  .nav-link a {{ color:var(--accent); text-decoration:none; font-size:0.85em; border:1px solid var(--accent); border-radius:20px; padding:4px 14px; }}
  .nav-link a:hover {{ background:var(--accent); color:var(--bg); }}
  .ace-explain {{ background:var(--box); border-radius:8px; padding:10px 14px; margin-bottom:14px; font-size:0.85em; }}
  .ace-explain summary {{ color:var(--accent); cursor:pointer; list-style:none; display:flex; align-items:center; gap:6px; min-height:44px; }}
  .ace-explain summary::-webkit-details-marker {{ display:none; }}
  .ace-explain summary::before {{ content:'ℹ'; font-size:1.1em; }}
  .ace-explain-hint {{ color:var(--muted); font-size:0.85em; }}
  .ace-explain p {{ color:var(--text); line-height:1.6; margin-top:8px; padding-top:8px; border-top:1px solid var(--border); }}
  .toggle {{ display:flex; justify-content:center; gap:8px; margin-bottom:16px; }}
  .toggle button {{ padding:8px 20px; border:1px solid var(--accent); background:transparent; color:var(--accent); border-radius:20px; cursor:pointer; font-size:0.9em; }}
  .toggle button.active {{ background:var(--accent); color:var(--bg); font-weight:bold; }}
  .basin-card {{ background:var(--card); border-radius:12px; padding:16px; margin-bottom:16px; display:none; }}
  .basin-card.active {{ display:block; }}
  h2 {{ color:var(--accent); font-size:1.2em; margin-bottom:6px; border-bottom:1px solid var(--border); padding-bottom:8px; }}
  .season-note {{ color:var(--muted); font-size:0.78em; margin-bottom:12px; }}
  .table-wrap {{ overflow-x:auto; background: linear-gradient(to right,var(--card) 20px,transparent 20px) left/20px 100%, linear-gradient(to left,var(--card) 20px,transparent 20px) right/20px 100%, linear-gradient(to right,rgba(0,0,0,0.18),transparent) left/16px 100%, linear-gradient(to left,rgba(0,0,0,0.18),transparent) right/16px 100%; background-repeat:no-repeat; background-attachment:local,local,scroll,scroll; }}
  table {{ width:100%; border-collapse:collapse; font-size:0.85em; }}
  th {{ background:var(--box); color:var(--accent); padding:8px 6px; text-align:left; position:sticky; top:0; white-space:nowrap; }}
  th.sort-th {{ cursor:pointer; user-select:none; padding:10px 6px; white-space:nowrap; }}
  th.sort-th:hover {{ color:var(--text-strong); }}
  .sa {{ font-size:0.7em; margin-left:2px; opacity:0.7; }}
  .hist-table th:first-child, .hist-table td:first-child {{ position:sticky; left:0; z-index:1; background:var(--box); border-right:1px solid var(--border); }}
  .hist-table tbody td:first-child {{ background:var(--card); }}
  .row-top5 {{ border-left:3px solid #f9a825; }}
  .row-top5 td:first-child {{ background:var(--card); }}
  .row-avg {{ border-top:2px solid var(--border); font-style:italic; }}
  .row-avg td {{ color:var(--muted); }}
  .row-avg td:first-child {{ background:var(--box); }}
  .ace-bar {{ height:3px; background:var(--gauge-bg); border-radius:2px; margin-top:3px; }}
  .ace-bar-fill {{ height:100%; background:var(--accent); border-radius:2px; }}
  .legend {{ display:flex; flex-wrap:wrap; gap:6px; justify-content:center; margin-bottom:14px; padding:10px; background:var(--card); border-radius:8px; }}
  .legend .badge {{ font-size:0.8em; padding:3px 10px; }}
  td {{ padding:7px 6px; border-bottom:1px solid var(--border); color:var(--text); white-space:nowrap; }}
  tr:hover td {{ filter:brightness(1.12); }}
  .row-extreme {{ background:var(--row-extreme); }}
  .row-above {{ background:var(--row-above); }}
  .row-near {{ background:var(--row-near); }}
  .row-below {{ background:var(--row-below); }}
  .row-current {{ border-left:3px solid var(--current-border); }}
  .badge {{ display:inline-block; padding:2px 8px; border-radius:12px; font-size:0.78em; font-weight:600; color:#fff; white-space:nowrap; }}
  .badge-extreme {{ background:var(--badge-extreme); }}
  .badge-above {{ background:var(--badge-above); }}
  .badge-near {{ background:var(--badge-near); }}
  .badge-below {{ background:var(--badge-below); }}
  .active-dot {{ color:var(--active-dot); font-size:0.65em; vertical-align:middle; margin-left:3px; }}
  .sources {{ background:var(--sources-bg); border-top:1px solid var(--border); margin-top:24px; padding:16px 12px; border-radius:8px; }}
  .sources h4 {{ color:var(--muted); font-size:0.8em; text-transform:uppercase; margin-bottom:8px; }}
  .sources a {{ color:var(--accent); text-decoration:none; font-size:0.78em; }}
  .sources a:hover {{ text-decoration:underline; }}
  .sources p {{ color:var(--muted-dark,#546e7a); font-size:0.75em; margin-top:8px; line-height:1.5; }}
  .sources ul {{ list-style:none; padding:0; margin:0; }}
  .sources li {{ color:var(--muted); font-size:0.78em; margin:4px 0; padding-left:12px; position:relative; }}
  .sources li::before {{ content:"•"; position:absolute; left:0; color:var(--accent); }}
  .sources code {{ font-size:0.9em; background:var(--box); padding:1px 4px; border-radius:3px; }}
  .disclaimer {{ margin-top:12px; padding:10px 12px; border-radius:6px; border-left:3px solid var(--muted); font-size:0.75em; color:var(--muted); line-height:1.5; }}
  .kofi-link {{ text-align:center; margin-top:14px; font-size:0.78em; }}
  .kofi-link a {{ color:var(--muted); text-decoration:none; }}
  .kofi-link a:hover {{ color:var(--accent); }}
  @media(min-width:768px) {{ body {{ max-width:960px; margin:0 auto; padding:24px; }} }}
  @media(min-width:1100px) {{ body {{ max-width:1280px; }} }}
</style>
</head>
<body>
<div class="header">
  <h1><img src="ace.png" class="logo" alt="ACE"> Hurricane ACE History</h1>
  <button class="theme-btn" id="themeBtn" onclick="toggleTheme()">☀</button>
</div>
<div class="updated">Updated: {now.strftime('%B %d, %Y at %H:%M UTC')}</div>
<div class="nav-link"><a href="index.html">← Current Season</a></div>
<details class="ace-explain">
  <summary>What is ACE? <span class="ace-explain-hint">(tap to expand)</span></summary>
  <p>Accumulated Cyclone Energy (ACE) measures total hurricane season activity by combining storm intensity and duration. A major hurricane that lasts two weeks contributes far more than a brief tropical storm. NOAA uses seasonal ACE totals to classify years as <b>Below Normal</b> (&lt;73), <b>Near Normal</b> (73–126), <b>Above Normal</b> (126–159), or <b>Extremely Active</b> (159+).</p>
</details>
<div class="toggle">
  <button class="active" onclick="show('atlantic',this)">Atlantic</button>
  <button onclick="show('pacific',this)">Eastern Pacific</button>
</div>
<div class="legend">
  <span class="badge badge-extreme">Extremely Active ≥159</span>
  <span class="badge badge-above">Above Normal 126–159</span>
  <span class="badge badge-near">Near Normal 73–126</span>
  <span class="badge badge-below">Below Normal &lt;73</span>
</div>
{''.join(basin_sections)}
<div class="sources">
  <h4>Data Sources</h4>
  <ul>
    <li><a href="https://www.nhc.noaa.gov/data/#hurdat" target="_blank" rel="noopener noreferrer">NOAA HURDAT2</a> — Official historical best-track database (1991–present) for all storm tracks, wind speeds, and ACE calculations</li>
    <li><a href="https://www.nhc.noaa.gov/data/#hurdat" target="_blank" rel="noopener noreferrer">NHC Real-time Best Track</a> — Current season preliminary storm data fetched via Tropycal (<code>include_btk=True</code>); updated continuously during active storms</li>
    <li><a href="https://www.cpc.ncep.noaa.gov/products/outlooks/background_information.shtml" target="_blank" rel="noopener noreferrer">NOAA CPC</a> — Season classification thresholds and 1991–2020 climatological normals</li>
  </ul>
  <p>ACE (Accumulated Cyclone Energy) is calculated at 6-hourly synoptic times (0000/0600/1200/1800 UTC) for systems at tropical storm strength or higher (≥34 kt), including subtropical storms. Formula: ACE = Σ(V²<sub>max</sub>) × 10⁻⁴. Categories use the Saffir-Simpson scale in knots.</p>
  <p class="disclaimer">⚠️ This site is maintained by a hurricane data enthusiast — not a meteorologist, forecaster, or weather professional of any kind. I just love the data. All information is sourced directly from official NOAA/NHC databases. For official forecasts, watches, warnings, and life-safety information, always refer to the <a href="https://www.nhc.noaa.gov/" target="_blank" rel="noopener noreferrer">National Hurricane Center</a>.</p>
  <p class="kofi-link"><a href="https://ko-fi.com/aceofcanes" target="_blank" rel="noopener noreferrer">☕ Support this project on Ko-fi</a></p>
</div>
<script>
document.querySelectorAll('.basin-card')[0]?.classList.add('active');
function show(id,btn) {{
  document.querySelectorAll('.basin-card').forEach(c=>c.classList.remove('active'));
  document.querySelectorAll('.toggle button').forEach(b=>b.classList.remove('active'));
  document.getElementById(id)?.classList.add('active');
  btn.classList.add('active');
  try{{history.replaceState(null,'','#'+id);}}catch(e){{}}
}}
function toggleTheme() {{
  var h=document.documentElement;
  var light=h.getAttribute('data-theme')==='light';
  h.setAttribute('data-theme',light?'dark':'light');
  try{{localStorage.setItem('ace-theme',light?'dark':'light');}}catch(e){{}}
  document.getElementById('themeBtn').textContent=light?'☀':'☾';
}}
document.addEventListener('DOMContentLoaded',function() {{
  document.getElementById('themeBtn').textContent=document.documentElement.getAttribute('data-theme')==='light'?'☾':'☀';
  var hash=location.hash.replace('#','');
  var match=[].slice.call(document.querySelectorAll('.toggle button')).filter(function(b){{return(b.getAttribute('onclick')||'').indexOf("'"+hash+"'")>=0;}})[0];
  if(match)match.click();
}});
var _hs={{}};
function sortHist(th,col,type){{
  var card=th.closest('.basin-card');
  var tbody=card.querySelector('tbody');
  var key=card.id+col;
  var asc=_hs[key]===undefined?false:!_hs[key];
  _hs[key]=asc;
  var rows=Array.from(tbody.querySelectorAll('tr'));
  rows.sort(function(a,b){{
    var av=a.cells[col]?a.cells[col].getAttribute('data-v'):'';
    var bv=b.cells[col]?b.cells[col].getAttribute('data-v'):'';
    if(type==='n'){{av=parseFloat(av)||0;bv=parseFloat(bv)||0;}}
    if(av<bv)return asc?-1:1;
    if(av>bv)return asc?1:-1;
    return 0;
  }});
  rows.forEach(r=>tbody.appendChild(r));
  card.querySelectorAll('.sort-th .sa').forEach((s,i)=>{{s.textContent=i===col?(asc?'▲':'▼'):'';}});
}}
</script>
</body>
</html>'''
    return html


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("\n" + "#" * 50)
    print("#" + " " * 48 + "#")
    print("#   HURRICANE ACE TRACKER - Atlantic & Pacific   #")
    print("#" + " " * 48 + "#")
    print("#" * 50)
    print(f"\nHistorical data from {START_YEAR} onward")
    print(f"ACE includes: TS + HU + SS statuses at synoptic times only")

    output_files = []
    basin_results = []

    # Process Atlantic
    result = process_basin('atlantic')
    if result:
        output_files.append(os.path.join(OUTPUT_FOLDER, BASINS['atlantic']['output_file']))
        basin_results.append(result)

    # Process Pacific
    result = process_basin('pacific')
    if result:
        output_files.append(os.path.join(OUTPUT_FOLDER, BASINS['pacific']['output_file']))
        basin_results.append(result)

    # Generate HTML pages
    if basin_results:
        dashboard_html = generate_dashboard_html(basin_results)
        dashboard_path = os.path.join(OUTPUT_FOLDER, 'ACE_Dashboard.html')
        try:
            with open(dashboard_path, 'w', encoding='utf-8') as f:
                f.write(dashboard_html)
            output_files.append(dashboard_path)
            print(f"\n  ✓ Dashboard saved to: {dashboard_path}")
        except (OSError, PermissionError) as e:
            logger.error(f"Failed to save dashboard {dashboard_path}: {e}")
            print(f"\n  ✗ Error: Could not save dashboard")

        history_html = generate_history_html(basin_results)
        history_path = os.path.join(OUTPUT_FOLDER, 'history.html')
        try:
            with open(history_path, 'w', encoding='utf-8') as f:
                f.write(history_html)
            output_files.append(history_path)
            print(f"  ✓ History page saved to: {history_path}")
        except (OSError, PermissionError) as e:
            logger.error(f"Failed to save history page {history_path}: {e}")
            print(f"  ✗ Error: Could not save history page")

    print("\n" + "=" * 50)
    print("All done! Spreadsheets and dashboard updated.")
    if output_files:
        print(f"Files saved to: {OUTPUT_FOLDER}")
        for f in output_files:
            print(f"  → {os.path.basename(f)}")
    print("=" * 50)

    return output_files


if __name__ == "__main__":
    main()
